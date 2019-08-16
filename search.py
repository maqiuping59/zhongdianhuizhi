import openpyxl  # 导入用来操作xlsx的库
import jieba  # 导入用来分词的库
from sys import argv
'''
    获取用户问题及相关的文件
'''
# row_index 用来表示行的索引
row_index = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
             'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
stopwords = ['你好', '请问', '，', '的', '-', ' 需要', '您好', '了解', '应该', '准备', '是不是', '不能',
             '没有', '>', '找到', '成都市', ' 可否', '？', '我', '没有', '一直', '都在', '吗', '我能', '都', '，',
             '尊敬', '本人', '_', '在', '年', '是', ' ', '将', '说', '用', '\n', '今年', '月', '需要', '日',
             '重庆', '2018', '给', '某', '1', '2', '3', '4', '5', '6', '7', '8', '没', '以后', '了',
             '刷', '深圳', '成华区', '哪些', '！', '能', '。']
# 导入训练集文件
wb1 = openpyxl.load_workbook(r"data\训练集.xlsx")
# 获取活跃的sheet
ws1 = wb1.active
# 导入公文条例集
wb = openpyxl.load_workbook(r"data\公文条例集.xlsx")
# 获取活跃的sheet

ws = wb.active

doc_dict = {}  # 定义一个字典用来保存文件

'''
    先创建一个字典来存储政策条例，键为公文名，值为条例名称
'''
for i in range(2, ws.max_row + 1):
    doc_dict[ws['B' + str(i)].value] = {}

for i in range(2, ws.max_row + 1):
    doc_dict[ws['B' + str(i)].value][ws['C' + str(i)].value] = {}

for i in range(2, ws.max_row + 1):
    doc_dict[ws['B' + str(i)].value][ws['C' + str(i)].value]['content'] = ws['D' + str(i)].value
    doc_dict[ws['B' + str(i)].value][ws['C' + str(i)].value]['feature'] = \
        [list(doc_dict.keys()).index(ws['B' + str(i)].value),
         list(doc_dict[ws['B' + str(i)].value].keys()).index(ws['C' + str(i)].value), {}]


def count_num(jieba_result, count_dict):
    '''
    输入为对问题分词后的列表，和用来记录每个词语对应该条例所出现的次数
    :param jieba_result:
    :param count_dict:
    :return:
    '''
    for item in jieba_result:
        count_dict[item] = count_dict.get(item, 0) + 1
    return count_dict


def get_distance(array=[], word1='', word2=''):
    '''
    获取输入字符串中的任意两词语之间的距离
    :param array:
    :param word1:
    :param word2:
    :return:
    '''
    if word1 and word2 in array:
        distance_1 = array.index(word1)
        distance_2 = array.index(word2)
        distance = np.abs(distance_1-distance_2)
        return distance
    else:
        return None


def find_doc(a):
    '''
    根据所索引可以查找对应的文件，参数为列表
    :param a:
    :return:
    '''
    dic1 = doc_dict[list(doc_dict.keys())[a[0]]]
    list1 = list(dic1.keys())
    dict2 = dic1[list1[0]]
    result = dict2['content']
    return result


def update_feature(feature_list, update_dict):
    '''
    特征提取之后更新公文条例的feature
    :param feature_list:
    :param update_dict:
    :return:
    '''
    result1 = doc_dict.get(list(doc_dict.keys())[feature_list[0]])
    result2 = result1.get(list(result1.keys())[feature_list[1]])
    result2['feature'][2].update(update_dict)
    # print(result2['feature'])


def function(a, b):
    '''
    根据文件名和条例名查询内容
    :param a:
    :param b:
    :return:
    '''
    a, b = str(a), str(b)
    a = a.lstrip()
    a = a.rstrip()
    result1 = doc_dict.get(a)
    if result1 is None:
        a = remove_kuohao(a)
        result1 = doc_dict.get(a)
    if type(result1) is dict:
        if type(b) is int:
                result2 = result1.get(b)
        else:
            result2 = result1.get(b)
            if result2 is None:
                result2 = result1.get(b.strip(' '))
                if result2 is None:
                    for items in result1:
                        content = result1.get(items)['content']
                        if b in content:
                            result2 = result1.get(items)
        return result2['feature']
    else:
        return None


def remove_kuohao(input_str=''):
    '''
    因为有些条例加了书名号，有些没加，所以该函数是为了去除书名号，便于查询
    :param input_str:
    :return:inputz_str
    '''
    if input_str.startswith('《'):
        return input_str[1:len(input_str)-1]


for i in range(2, ws1.max_row):
    question = ws1["B"+str(i)].value
    if question is None:
        break
    words = jieba.lcut(question)
    for j in range(0, len(row_index), 2):
        if ws1[row_index[j]+str(i)].value is not None:
            doc = ws1[row_index[j]+str(i)].value       # 获取文件名
            item = ws1[row_index[j+1]+str(i)].value    # 获取条例名称
            try:
                result = function(doc, item)
                pre_dict = result[2]
                for k in words:
                    pre_dict[k] = pre_dict.get(k, 0)+1
                update_feature(result, pre_dict)
                pre_dict = {}
            except TypeError:
                pass
        else:
            break


def search(input_str):
    '''
    根据问题查询对应的公文条例
    :param input_str:
    :return:
    '''
    seg_result = jieba.lcut(input_str)
    answer_ = '|'
    answer_file = '您好！'
    max_pos = 0
    pre_result = []
    seg_result = set(seg_result)
    for sin in list(seg_result):
        if sin in stopwords:
            try:
                seg_result.remove(sin)
            except KeyError:
                pass
    for feature in feature_matrix:
        reference = 0
        common_part = []
        for word in seg_result:
            reference = feature[2].get(word, 0) + reference
            if word in feature[2].keys():
                common_part.append(word)

        if reference > 0.1:
            if reference > max_pos:
                max_pos = reference
            pre_result.append([feature, reference])
    for feature_ in pre_result:
        if feature_[1] > 0.9*max_pos:
            result_ = search_title(feature_[0])
            answer_ = answer_ + result_[0]+'|'+result_[1]+'|'
            regular = "根据"+'《'+result_[0]+'》'+"的相关规定"
            answer_file = answer_file+regular+doc_dict[result_[0]][result_[1]]['content']
    return answer_file+answer_


def search_title(a):
    '''
    根据每个条例下的特征查找其所对应的公文名称及条例名称
    :param a:
    :return:
    '''
    result1 = list(doc_dict.keys())[a[0]]
    result2 = list(doc_dict.get(result1).keys())[a[1]]
    return result1, result2


if __name__ == "__main__":
    feature_matrix = []
    matrix_row = 0
    for m in doc_dict:
        one = doc_dict.get(m)
        for n in one:
            two = one.get(n)['feature'][2]
            for o in list(two.keys()):
                if o in stopwords:
                    try:
                        del two[o]
                    except KeyError:
                        pass
            add_all = sum(two.values())
            for item in two.keys():
                two[item] = two.get(item) / add_all
            # print(two)
            doc_dict[m][n]['feature'][2] = two
            feature_matrix.append(doc_dict[m][n]['feature'])
            matrix_row = matrix_row + 1
    wb1.close()
    wb.close()
    test = argv[1]
    with open('result.txt', 'w', encoding='utf-8') as res:
        with open(test, 'r') as que:
            questions = que.readlines()
            for question in questions:
                if question is not None:
                    # print(question)
                    answer = search(question)
                    # print(answer)
                    res.writelines(answer)
                    res.write('\n\n')
                else:
                    break


